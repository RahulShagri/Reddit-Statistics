import praw
from datetime import datetime
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils import get_column_letter

workbook = openpyxl.load_workbook('input.xlsx')
workbook_out = openpyxl.load_workbook('output.xlsx')
sheet = workbook['Sheet1']
sheet_out = workbook_out['Sheet1']

sheet_out['A1'].value = 'Number of Posts'
sheet_out['B1'].value = 'Time'

reddit = praw.Reddit(client_id='your client id',
                     client_secret='your client secret',
                     user_agent='your user agent',
                     username='your username',
                     password='your password')

data = []

for i in range(50):

    sub = sheet['A'+str(i+2)].value
    subreddit = reddit.subreddit(sub)

    for submission in subreddit.top('all', limit = 100):
        temp = int(submission.created_utc)
        data.append(int(datetime.utcfromtimestamp(temp).strftime('%H')))

    print(str(i+1)+' Subreddit data collected!')

posts = []
hour = []
for i in range(24):
    hour.append(str(i)+':00 - '+str(i+1)+':00')
    if i in data:
        posts.append(data.count(i))
    else:
        posts.append(0)
    sheet_out['A'+str(i+2)].value = posts[i]
    sheet_out['B'+str(i+2)].value = hour[i]

workbook_out.save('output.xlsx')
print('Output written to file!')

plt.bar(hour, posts)
plt.ylabel('Total number of posts')
plt.xlabel('Time')
plt.xticks(rotation = 40)
plt.show()
